#!/usr/bin/env python
"""Generate the synthetic demo dataset (demo/) for cashato.

Produces realistic — but entirely fictional — bank statements in every format
the three adapters support, from ONE deterministic ground truth (seeded RNG):

- Intesa Sanpaolo: 6 quarterly statement PDFs (Italian layout, double date,
  debit/credit columns) + a 13-month "Lista Operazioni" XLSX that OVERLAPS the
  quarterlies (cross-file dedup test).
- Revolut: consolidated-statement CSV and PDF carrying the SAME movements
  (cross-format dedup test), two accounts (Personal + Joint), savings-interest
  and crypto-sales sections (CSV only, like the real parser expects).
- Trade Republic: statement PDF (Italian, position-aware columns) and the
  transaction-export CSV, where amount+fee+tax in the CSV equals the PDF's net
  amount (cross-format dedup test).

Deliberate edge cases baked in:
- two identical same-day POS payments (occurrence-index disambiguation);
- monthly internal transfers Intesa->Revolut and Intesa->Trade Republic
  (opposite legs within the transfer window -> transfer_group pairing);
- a Revolut ATM withdrawal with a fee (separate fee transaction);
- value date != booking date on some Intesa rows.

The PDF layouts are generated to the SAME geometry the parsers are calibrated
on (word X coordinates), because parsing is position-aware. The persona is
fictional: MARIO BIANCHI, Via Garibaldi 42, 20121 Milano.

Usage:  .venv/bin/python demo/generate.py [--out demo]
Requires: fpdf2 (dev-only), openpyxl (already a project dependency).
"""

from __future__ import annotations

import argparse
import csv
import random
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

from fpdf import FPDF

# --- persona (fictional!) -----------------------------------------------------

HOLDER_GIVEN_FIRST = "MARIO BIANCHI"  # Revolut / Trade Republic order
HOLDER_FAMILY_FIRST = "BIANCHI MARIO"  # Italian bank statements order
STREET = "Via Garibaldi 42"
CAP_CITY = "20121 Milano"
# Structurally valid Italian IBAN (check digits fake), ABI 03069 = Intesa
IBAN_INTESA = "IT60X0306901000100000012345"
IBAN_REVOLUT = "LT313250048123456789"  # not Italian: find_iban() ignores it

START = date(2025, 1, 1)
END = date(2026, 6, 30)
XLSX_START = date(2025, 6, 1)  # 13-month export window (overlaps quarterlies)
SEED = 42

D0 = Decimal("0.00")


def eur(x: str | int | float) -> Decimal:
    return Decimal(str(x)).quantize(Decimal("0.01"))


def fmt_it(v: Decimal) -> str:
    """1234.56 -> '1.234,56' (unsigned)."""
    s = f"{abs(v):,.2f}"
    return s.replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_en(v: Decimal, *, signed: bool = True) -> str:
    """1234.56 -> '1,234.56'; signed adds a leading '-' for outflows."""
    s = f"{abs(v):,.2f}"
    return f"-{s}" if signed and v < 0 else s


def months(start: date, end: date):
    d = start.replace(day=1)
    while d <= end:
        yield d
        d = (d + timedelta(days=32)).replace(day=1)


# --- ground truth -------------------------------------------------------------


@dataclass
class IMov:  # Intesa movement
    booking: date
    value: date
    amount: Decimal  # signed
    line1: str  # first description line (PDF)
    line2: str = ""  # continuation line (PDF)
    op: str = ""  # XLSX "Operazione"
    det: str = ""  # XLSX "Dettagli"
    cat: str = ""  # XLSX "Categoria" (native, bootstrap-only)


@dataclass
class RMov:  # Revolut cash movement (same in CSV and PDF)
    d: date
    desc: str
    cat: str
    amount: Decimal  # signed
    fee: Decimal = D0


@dataclass
class TMov:  # Trade Republic movement
    d: date
    tipo: str  # PDF "TIPO" column (Italian)
    desc: str  # PDF description
    net: Decimal  # signed net cash impact (= PDF amount = csv amount+fee+tax)
    csv_type: str = ""
    name: str = ""
    desc_csv: str = ""
    amount: Decimal = D0  # CSV amount (gross)
    fee: Decimal = D0
    tax: Decimal = D0
    cat: str = ""
    mcc: str = ""
    asset: str = ""
    isin: str = ""


@dataclass
class GroundTruth:
    intesa: list[IMov] = field(default_factory=list)
    rev: dict[str, list[RMov]] = field(default_factory=lambda: {"Personal": [], "Joint": []})
    rev_interest: list[tuple[date, Decimal]] = field(default_factory=list)  # (date, net)
    tr: list[TMov] = field(default_factory=list)


_GROCERIES = ["ESSELUNGA MILANO", "CARREFOUR MARKET MILANO", "LIDL MILANO VIA PADOVA"]
_FUEL = ["Q8 MILANO V.LE MONZA", "ENI STATION MILANO EST"]
_RESTAURANTS_IT = ["RISTORANTE DA GINO MILANO", "PIZZERIA BELLA NAPOLI MILANO"]
_REV_FOOD = [("Lidl", "Groceries"), ("Penny Market", "Groceries"), ("Carrefour Express", "Groceries")]
_REV_OUT = [("Pizzeria Spontini", "Restaurants"), ("Sushi Daily", "Restaurants"), ("Bar Luce", "Cafes")]
_REV_TRANSPORT = [("ATM Milano", "Transport"), ("Trenord", "Transport")]
_TR_CARD = [("REWE Berlin", "5411"), ("Esselunga Milano", "5411"), ("Amazon EU", "5942")]


def _ramt(rng: random.Random, lo: float, hi: float) -> Decimal:
    return eur(round(rng.uniform(lo, hi), 2))


def build_ground_truth(rng: random.Random) -> GroundTruth:
    gt = GroundTruth()
    for m0 in months(START, END):
        y, m = m0.year, m0.month

        def day(d: int, y: int = y, m: int = m) -> date:
            return date(y, m, min(d, 28))

        # --- Intesa: the salary account -------------------------------------
        gt.intesa.append(
            IMov(day(27), day(27), eur(2450) + _ramt(rng, 0, 40),
                 "BONIFICO A VOSTRO FAVORE DA ACME SRL",
                 "ACCREDITO EMOLUMENTI", "Accredito stipendio", "Acme Srl emolumenti", "Stipendio"))
        gt.intesa.append(
            IMov(day(1), day(1), -eur(850),
                 "BONIFICO SEPA A ROSSI IMMOBILIARE",
                 "CANONE LOCAZIONE", "Bonifico disposto", "Rossi Immobiliare canone locazione", "Casa"))
        gt.intesa.append(
            IMov(day(16), day(16), -eur("29.90"),
                 "ADDEBITO SDD TIM SPA",
                 "TELEFONO FISSO E FIBRA", "Addebito SDD", "Tim Spa telefono e fibra", "Utenze"))
        if m % 2 == 0:
            gt.intesa.append(
                IMov(day(12), day(12), -_ramt(rng, 55, 95),
                     "ADDEBITO SDD ENEL ENERGIA SPA",
                     "FORNITURA ENERGIA ELETTRICA", "Addebito SDD", "Enel Energia fornitura", "Utenze"))
        for _ in range(rng.randint(3, 5)):
            merch = rng.choice(_GROCERIES)
            b = day(rng.randint(2, 26))
            v = b + timedelta(days=rng.choice([0, 0, 0, 1, 2]))
            gt.intesa.append(
                IMov(b, v, -_ramt(rng, 18, 95),
                     "PAGAMENTO POS CARTA *4523", merch, "Pagamento POS", merch.title(), "Spesa"))
        for _ in range(rng.randint(1, 2)):
            merch = rng.choice(_FUEL)
            b = day(rng.randint(3, 25))
            gt.intesa.append(
                IMov(b, b, -_ramt(rng, 38, 70),
                     "PAGAMENTO POS CARTA *4523", merch, "Pagamento POS", merch.title(), "Auto"))
        if rng.random() < 0.7:
            merch = rng.choice(_RESTAURANTS_IT)
            b = day(rng.randint(5, 27))
            gt.intesa.append(
                IMov(b, b, -_ramt(rng, 25, 80),
                     "PAGAMENTO POS CARTA *4523", merch, "Pagamento POS", merch.title(), "Ristoranti"))
        if rng.random() < 0.6:
            b = day(rng.randint(4, 24))
            gt.intesa.append(
                IMov(b, b, -_ramt(rng, 9, 40),
                     "PAGAMENTO POS CARTA *4523", "FARMACIA CENTRALE MILANO",
                     "Pagamento POS", "Farmacia Centrale Milano", "Salute"))
        b = day(rng.randint(6, 22))
        gt.intesa.append(
            IMov(b, b, -eur(rng.choice([100, 150, 200])),
                 "PRELIEVO CARTA DEBITO SU ATM",
                 "MILANO VIA MANZONI", "Prelievo", "Prelievo ATM Milano", "Contanti"))
        if rng.random() < 0.8:
            b = day(rng.randint(2, 26))
            gt.intesa.append(
                IMov(b, b, -_ramt(rng, 12, 90),
                     "PAGAMENTO POS CARTA *4523", "AMAZON EU LUXEMBOURG",
                     "Pagamento POS", "Amazon Eu", "Shopping"))

        # --- internal transfers: Intesa -> Revolut, Intesa -> Trade Republic --
        topup = eur(300)
        b = day(5)
        gt.intesa.append(
            IMov(b, b, -topup,
                 "PAGAMENTO POS CARTA *4523", "RICARICA CARTA DUBLINO IRL",
                 "Pagamento POS", "Ricarica carta Dublino", "Trasferimenti"))
        gt.rev["Personal"].append(RMov(b + timedelta(days=1), "Top-Up by *4523", "Transfers", topup))

        deposit = eur(400)
        b = day(3)
        gt.intesa.append(
            IMov(b, b + timedelta(days=1), -deposit,
                 "BONIFICO EUROPEO SEPA DISPOSTO A", "FAVORE DI T.R. BANK GMBH BERLINO",
                 "Bonifico disposto", "T.R. Bank Gmbh Berlino", "Trasferimenti"))
        gt.tr.append(
            TMov(b + timedelta(days=1), "Bonifico", "Accredito bonifico da Mario Bianchi",
                 deposit, csv_type="DEPOSIT", name="Mario Bianchi",
                 desc_csv="Incoming bank transfer", amount=deposit, cat="transfer"))

        # --- Revolut Personal: subscriptions + daily spending -----------------
        gt.rev["Personal"].append(RMov(day(14), "Netflix.com", "Entertainment", -eur("12.99")))
        gt.rev["Personal"].append(RMov(day(20), "Spotify", "Entertainment", -eur("10.99")))
        for _ in range(rng.randint(2, 3)):
            merch, cat = rng.choice(_REV_FOOD)
            gt.rev["Personal"].append(
                RMov(day(rng.randint(2, 27)), merch, cat, -_ramt(rng, 8, 45)))
        for _ in range(rng.randint(1, 3)):
            merch, cat = rng.choice(_REV_OUT)
            gt.rev["Personal"].append(
                RMov(day(rng.randint(2, 27)), merch, cat, -_ramt(rng, 9, 55)))
        for _ in range(rng.randint(2, 4)):
            merch, cat = rng.choice(_REV_TRANSPORT)
            gt.rev["Personal"].append(
                RMov(day(rng.randint(2, 27)), merch, cat, -eur("2.20")))
        if rng.random() < 0.4:
            gt.rev["Personal"].append(
                RMov(day(rng.randint(3, 26)), "To Giulia Verdi", "Transfers", -_ramt(rng, 15, 60)))

        # --- Revolut Joint: shared expenses -----------------------------------
        for _ in range(rng.randint(2, 3)):
            merch, cat = rng.choice(_REV_FOOD + [("IKEA Carugate", "Shopping")])
            gt.rev["Joint"].append(
                RMov(day(rng.randint(2, 27)), merch, cat, -_ramt(rng, 15, 90)))

        # --- Revolut savings interest (CSV-only section) -----------------------
        gt.rev_interest.append((day(28), _ramt(rng, 1.05, 1.95)))

        # --- Trade Republic: ETF plan, interest, card --------------------------
        gt.tr.append(
            TMov(day(15), "Risparmio", "Acquisto ETF IE00B4L5Y983 Core MSCI World",
                 -eur(200), csv_type="TRADE", name="iShares Core MSCI World USD Acc",
                 desc_csv="Savings plan execution", amount=-eur(200),
                 cat="funds", asset="ETF", isin="IE00B4L5Y983"))
        if m % 3 == 0:
            gt.tr.append(
                TMov(day(9), "Acquisto", "Acquisto ETF IE00B3RBWM25 FTSE All-World",
                     -eur(501), csv_type="TRADE", name="Vanguard FTSE All-World UCITS",
                     desc_csv="Buy order", amount=-eur(500), fee=-eur(1),
                     cat="funds", asset="ETF", isin="IE00B3RBWM25"))
        gross = _ramt(rng, 1.60, 3.40)
        tax = -eur(round(float(gross) * 0.26, 2))
        gt.tr.append(
            TMov(day(28), "Interessi", "Interessi maturati sulla liquidita",
                 gross + tax, csv_type="INTEREST", name="Trade Republic",
                 desc_csv="Interest payout", amount=gross, tax=tax, cat="interest"))
        for _ in range(rng.randint(1, 2)):
            merch, mcc = rng.choice(_TR_CARD)
            amt = -_ramt(rng, 10, 70)
            gt.tr.append(
                TMov(day(rng.randint(2, 27)), "Carta", merch, amt,
                     csv_type="CARD_PAYMENT", name=merch, desc_csv="Card payment",
                     amount=amt, cat="card", mcc=mcc))

    # --- deliberate edge cases -------------------------------------------------
    # 1) two IDENTICAL same-day POS payments (occurrence-index disambiguation)
    twin = date(2025, 3, 14)
    for _ in range(2):
        gt.intesa.append(
            IMov(twin, twin, -eur("1.20"),
                 "PAGAMENTO POS CARTA *4523", "BAR CENTRALE MILANO",
                 "Pagamento POS", "Bar Centrale Milano", "Ristoranti"))
    # 2) an ATM withdrawal with a fee on Revolut (separate fee transaction)
    gt.rev["Personal"].append(
        RMov(date(2025, 8, 9), "ATM Withdrawal Lisboa", "Cash", -eur(60), fee=eur(2)))

    for movs in (gt.intesa, gt.tr, *gt.rev.values()):
        movs.sort(key=lambda t: (t.booking if isinstance(t, IMov) else t.d))
    return gt


# --- PDF helpers ----------------------------------------------------------------


def new_pdf() -> FPDF:
    pdf = FPDF(unit="pt", format="A4")
    pdf.set_auto_page_break(False)
    pdf.set_text_color(0, 0, 0)
    return pdf


def put(pdf: FPDF, x: float, y: float, text: str, size: float = 8, style: str = "",
        max_x: float | None = None) -> None:
    pdf.set_font("Helvetica", style, size)
    if max_x is not None and x + pdf.get_string_width(text) > max_x:
        raise ValueError(f"text overflows column ({x:.0f}->{max_x:.0f}): {text!r}")
    pdf.text(x, y, text)


def put_right(pdf: FPDF, x1: float, y: float, text: str, size: float = 8, style: str = "") -> None:
    pdf.set_font("Helvetica", style, size)
    pdf.text(x1 - pdf.get_string_width(text), y, text)


def addressee_block(pdf: FPDF, x: float, y: float, name: str) -> None:
    """Name / street / CAP+city — the layout base.addressee_from_words expects."""
    put(pdf, x, y, name, size=9)
    put(pdf, x, y + 12, STREET, size=9)
    put(pdf, x, y + 24, CAP_CITY, size=9)


# --- Intesa quarterly PDFs -------------------------------------------------------

# Geometry the parser is calibrated on: booking date x<70, value date 85..150,
# description 150..360, amounts x>=360, debit/credit split at Accrediti.x0-30.
_I_X_BOOK, _I_X_VAL, _I_X_DESC = 38, 88, 152
_I_X1_DEBIT, _I_X0_CREDIT = 450, 490
_I_X1_CREDIT = 545


def _intesa_header_row(pdf: FPDF, y: float) -> None:
    put(pdf, _I_X_BOOK, y, "Data Oper.", size=8, style="B")
    put(pdf, _I_X_VAL, y, "Data Valuta", size=8, style="B")
    put(pdf, _I_X_DESC, y, "Descrizione", size=8, style="B")
    put_right(pdf, _I_X1_DEBIT, y, "Addebiti", size=8, style="B")
    put(pdf, _I_X0_CREDIT, y, "Accrediti", size=8, style="B")


def write_intesa_quarter(path: Path, movs: list[IMov], n: int, year: int,
                         period: tuple[date, date], opening: Decimal) -> Decimal:
    pdf = new_pdf()
    # page 1: letterhead, product, IBAN, addressee — no movements table
    pdf.add_page()
    put(pdf, 40, 52, "Intesa Sanpaolo S.p.A. - Sede Legale Torino", size=11, style="B")
    put(pdf, 40, 82, f"ESTRATTO CONTO N. {n}/{year}", size=10, style="B")
    put(pdf, 40, 100, f"Periodo dal {period[0]:%d.%m.%Y} al {period[1]:%d.%m.%Y}", size=9)
    put(pdf, 40, 130, "Tipologia conto:", size=9)
    put(pdf, 40, 144, "XME Conto", size=9)
    put(pdf, 40, 172, f"IBAN: {IBAN_INTESA}", size=9)
    addressee_block(pdf, 360, 90, HOLDER_FAMILY_FIRST)
    closing = opening + sum(t.amount for t in movs)
    put(pdf, 40, 210, f"Saldo iniziale: {fmt_it(opening)}  Saldo finale: {fmt_it(closing)}", size=9)

    y = 1e9  # force a new page before the first row
    first = True
    for t in movs:
        rows = 2 if t.line2 else 1
        if y + 13 * rows > 780:
            pdf.add_page()
            _intesa_header_row(pdf, 60)
            y = 80
            if first:
                # Like the real statement: "al" the LAST day of the previous
                # period (the balance BEFORE this quarter's movements), amount
                # SIGNED ("+3.434,25") — the sign is how a debit balance shows.
                put(pdf, _I_X_DESC, y,
                    f"Saldo iniziale al {period[0] - timedelta(days=1):%d.%m.%Y}", size=7.5)
                put_right(pdf, _I_X1_CREDIT, y,
                          ("-" if opening < 0 else "+") + fmt_it(opening), size=7.5)
                y += 13
                first = False
        put(pdf, _I_X_BOOK, y, f"{t.booking:%d.%m.%Y}", size=7.5)
        put(pdf, _I_X_VAL, y, f"{t.value:%d.%m.%Y}", size=7.5)
        put(pdf, _I_X_DESC, y, t.line1, size=7.5, max_x=358)
        if t.amount < 0:
            put_right(pdf, _I_X1_DEBIT, y, fmt_it(t.amount), size=7.5)
        else:
            put_right(pdf, _I_X1_CREDIT, y, fmt_it(t.amount), size=7.5)
        if t.line2:
            y += 11
            put(pdf, _I_X_DESC, y, t.line2, size=7.5, max_x=358)
        y += 13
    put(pdf, _I_X_DESC, y, f"Saldo finale al {period[1]:%d.%m.%Y}", size=7.5)
    put_right(pdf, _I_X1_CREDIT, y, ("-" if closing < 0 else "+") + fmt_it(closing), size=7.5)
    for i in range(1, pdf.pages_count + 1):
        pdf.page = i
        put(pdf, 40, 810, f"Pagina {i} di {pdf.pages_count}", size=7)
    pdf.output(str(path))
    return closing


def write_intesa_quarters(outdir: Path, movs: list[IMov]) -> list[Path]:
    paths = []
    opening = eur(3000)
    quarters = [(y, q) for y in (2025, 2026) for q in (1, 2, 3, 4)
                if date(y, 3 * q - 2, 1) <= END]
    for y, q in quarters:
        p0 = date(y, 3 * q - 2, 1)
        p1 = (p0 + timedelta(days=93)).replace(day=1) - timedelta(days=1)
        batch = [t for t in movs if p0 <= t.booking <= p1]
        path = outdir / f"intesa_estratto_conto_{y}_Q{q}.pdf"
        opening = write_intesa_quarter(path, batch, q, y, (p0, p1), opening)
        paths.append(path)
    return paths


# --- Intesa 13-month XLSX ---------------------------------------------------------


def write_intesa_xlsx(path: Path, movs: list[IMov]) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista Operazioni"
    ws.append(["Lista Operazioni"])
    ws.append([f"Intestatario: {HOLDER_FAMILY_FIRST.title()}"])
    ws.append([f"Periodo: {XLSX_START:%d.%m.%Y} - {END:%d.%m.%Y}"])
    ws.append([])
    # The real export opens with a filter recap; "Conti e Carte:" is the string
    # content detection anchors on (the export carries no IBAN and never names
    # the bank), so the demo file must carry it too.
    ws.append([None, "Conti e Carte:", "Conto 1000 / 00012345"])
    ws.append([None, "Finanziamento:", "-"])
    ws.append([])
    ws.append(["Data", "Operazione", "Dettagli", "Categoria", "Valuta", "Importo"])
    # The export lists by VALUE date (which is what the quarterlies' natural_key
    # uses too — that is exactly what makes the overlap dedup work).
    for t in sorted(movs, key=lambda t: t.value):
        if not (XLSX_START <= t.value <= END):
            continue
        ws.append([datetime(t.value.year, t.value.month, t.value.day),
                   t.op, t.det, t.cat, "EUR", float(t.amount)])
    wb.save(str(path))


# --- Revolut CSV -------------------------------------------------------------------

_REV_TX_HEADER = ["Date", "Description", "Category", "Money in/out", "Balance",
                  "Tax withheld", "Other taxes", "Fees"]


def _rev_date(d: date) -> str:
    return f"{d:%b} {d.day}, {d.year}"


def _rev_details_rows(iban: str, modality: str) -> list[list[str]]:
    return [
        ["Account number", iban],
        ["Financial institution name", "Revolut Bank UAB"],
        ["Holding modalities", modality],
    ]


def write_revolut_csv(path: Path, gt: GroundTruth, balances: dict[str, list[Decimal]]) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Current Accounts Summaries"])
        w.writerow([])
        for label in ("Personal", "Joint"):
            w.writerow([f"{label} Account (EUR)"])
            w.writerows(_rev_details_rows(IBAN_REVOLUT, "Individual" if label == "Personal" else "Joint"))
            w.writerow([])
            w.writerow([f"Transaction statement from {_rev_date(START)} to {_rev_date(END)}"])
            w.writerow(_REV_TX_HEADER)
            for t, bal in zip(gt.rev[label], balances[label], strict=True):
                fee = f"€{fmt_en(t.fee, signed=False)}" if t.fee else ""
                w.writerow([_rev_date(t.d), t.desc, t.cat,
                            fmt_en(t.amount), fmt_en(bal), "", "", fee])
            w.writerow(["---------"])
        # savings-interest section (parsed from the CSV only)
        w.writerow(["Savings Account (EUR)"])
        w.writerow(["This statement shows only interest receipt transactions"])
        w.writerow(["Date", "Description", "Gross rate", "Gross amount", "Tax",
                    "Tax rate", "Fees", "Net amount"])
        for d, net in gt.rev_interest:
            gross = net  # keep it simple: no tax on the demo savings account
            w.writerow([_rev_date(d), "Gross interest", "2.25%", fmt_en(gross),
                        "0.00", "0%", "0.00", fmt_en(net)])
        w.writerow(["---------"])
        # crypto-sales section (two sales, account revolut_crypto)
        w.writerow(["Crypto Statements (EUR)"])
        w.writerow(["This statement shows only sales transactions"])
        w.writerow(["Date (of Sale, of Purchase)", "Symbol", "Quantity",
                    "Price (of Sale, of Purchase)", "Fees", "Value (of Sale, of Purchase)"])
        w.writerow(["15/03/2026, 10/01/2026", "BTC", "0.0016",
                    "€93,750.00, €75,000.00", "€0.00", "+ €150.00, - €120.00"])
        w.writerow(["02/05/2026, 08/02/2026", "ETH", "0.05",
                    "€3,400.00, €2,900.00", "€0.00", "+ €170.00, - €145.00"])
        w.writerow(["---------"])


# --- Revolut PDF -------------------------------------------------------------------

# Parser column map (x0): date <120 | desc 120-257 | category 257-320 |
# money 320-380 | balance 380-425 | fees >=520.
_R_X_DATE, _R_X_DESC, _R_X_CAT, _R_X_MONEY, _R_X_BAL, _R_X_FEE = 40, 122, 259, 325, 385, 525


def _revolut_pdf_header(pdf: FPDF, y: float) -> None:
    put(pdf, _R_X_DATE, y, "Date", size=8, style="B")
    put(pdf, _R_X_DESC, y, "Description", size=8, style="B")
    put(pdf, _R_X_CAT, y, "Category", size=8, style="B")
    put(pdf, _R_X_MONEY, y, "Money in/out", size=8, style="B")
    put(pdf, _R_X_BAL, y, "Balance", size=8, style="B")
    put(pdf, _R_X_FEE, y, "Fees", size=8, style="B")


def write_revolut_pdf(path: Path, gt: GroundTruth, balances: dict[str, list[Decimal]]) -> None:
    pdf = new_pdf()
    pdf.add_page()
    put(pdf, 40, 52, "Revolut Bank UAB - Konstitucijos ave. 21B Vilnius", size=10, style="B")
    addressee_block(pdf, 40, 90, HOLDER_GIVEN_FIRST)
    put(pdf, 40, 140, f"Consolidated statement {_rev_date(START)} - {_rev_date(END)}", size=9)
    y = 170
    for label in ("Personal", "Joint"):
        if y > 700:
            pdf.add_page()
            y = 60
        put(pdf, 40, y, f"{label} Account (EUR)", size=9, style="B")
        y += 16
        _revolut_pdf_header(pdf, y)
        y += 14
        for t, bal in zip(gt.rev[label], balances[label], strict=True):
            if y > 790:
                pdf.add_page()
                y = 60
                put(pdf, 40, y, f"{label} Account (EUR)", size=9, style="B")
                y += 16
                _revolut_pdf_header(pdf, y)
                y += 14
            put(pdf, _R_X_DATE, y, _rev_date(t.d), size=8)
            put(pdf, _R_X_DESC, y, t.desc, size=8, max_x=255)
            put(pdf, _R_X_CAT, y, t.cat, size=8)
            put(pdf, _R_X_MONEY, y, fmt_en(t.amount), size=8)
            put(pdf, _R_X_BAL, y, fmt_en(bal), size=8)
            if t.fee:
                put(pdf, _R_X_FEE, y, fmt_en(t.fee, signed=False), size=8)
            y += 13
        y += 24
    pdf.output(str(path))


def revolut_balances(gt: GroundTruth) -> dict[str, list[Decimal]]:
    """Running balance per account, one entry per movement.

    The Fees column is informational — on real statements the balance moves by
    ``Money in/out`` alone (the fee is already inside it), so the running
    balance here must NOT subtract it again.
    """
    out: dict[str, list[Decimal]] = {}
    # Openings high enough that no account ever goes negative: Revolut accounts
    # cannot run an overdraft, so a statement with a negative balance would be
    # unrealistic (the joint account only ever spends in this dataset).
    for label, start in (("Personal", eur(350)), ("Joint", eur(2800))):
        bal, seq = start, []
        for t in gt.rev[label]:
            bal += t.amount
            seq.append(bal)
        out[label] = seq
    return out


# --- Trade Republic PDF -------------------------------------------------------------

_T_X_DATA, _T_X_TIPO, _T_X_DESC = 38, 100, 150
_T_X_ENT, _T_X_USC, _T_X_SAL = 345, 435, 520  # header x0; amounts right-align to x1


def write_tr_pdf(path: Path, movs: list[TMov]) -> None:
    pdf = new_pdf()
    pdf.add_page()
    put(pdf, 40, 52, "Trade Republic Bank GmbH - Brunnenstrasse 19-21, 10119 Berlin", size=10, style="B")
    addressee_block(pdf, 40, 92, HOLDER_GIVEN_FIRST)
    put(pdf, 400, 92, "RENDICONTO", size=10, style="B")
    put(pdf, 400, 106, f"{START:%d.%m.%Y} - {END:%d.%m.%Y}", size=8)
    put(pdf, 40, 150, "IBAN DE00 1101 0100 1234 5678 91", size=8)

    # header ONCE (the parser derives the column geometry from these words'
    # right edges; repeating it on later pages would leak into descriptions)
    y = 180.0
    put(pdf, _T_X_DATA, y, "DATA", size=8, style="B")
    put(pdf, _T_X_TIPO, y, "TIPO", size=8, style="B")
    put(pdf, _T_X_DESC, y, "DESCRIZIONE", size=8, style="B")
    put(pdf, _T_X_ENT - 14, y, "IN", size=8, style="B")
    put(pdf, _T_X_ENT, y, "ENTRATA", size=8, style="B")
    pdf.set_font("Helvetica", "B", 8)
    ent_x1 = _T_X_ENT + pdf.get_string_width("ENTRATA")
    put(pdf, _T_X_USC - 14, y, "IN", size=8, style="B")
    put(pdf, _T_X_USC, y, "USCITA", size=8, style="B")
    pdf.set_font("Helvetica", "B", 8)
    usc_x1 = _T_X_USC + pdf.get_string_width("USCITA")
    put(pdf, _T_X_SAL, y, "SALDO", size=8, style="B")
    pdf.set_font("Helvetica", "B", 8)
    sal_x1 = _T_X_SAL + pdf.get_string_width("SALDO")
    y += 16

    # Opening balance: a cash account at a broker does not run an overdraft, so
    # the statement's SALDO column never goes negative — start it high enough.
    bal = eur(500)
    # balance-only row: becomes a dateless anchor the parser discards, absorbing
    # the header words that would otherwise leak into the first row's description
    put(pdf, _T_X_DESC, y, "Saldo iniziale", size=7)
    put_right(pdf, sal_x1, y, fmt_it(bal), size=8)
    y += 14

    for t in movs:
        if y > 800:
            pdf.add_page()
            y = 60
        bal += t.net
        put(pdf, _T_X_DATA, y, f"{t.d.day:02d}", size=8)
        put(pdf, _T_X_DATA + 16, y, _MONTH_IT[t.d.month], size=8)
        put(pdf, _T_X_DATA + 36, y, str(t.d.year), size=8)
        put(pdf, _T_X_TIPO, y, t.tipo, size=7, max_x=148)
        put(pdf, _T_X_DESC, y, t.desc, size=7, max_x=int(ent_x1 - 47))
        if t.net >= 0:
            put_right(pdf, ent_x1, y, fmt_it(t.net), size=8)
        else:
            put_right(pdf, usc_x1, y, fmt_it(t.net), size=8)
        put_right(pdf, sal_x1, y, fmt_it(bal), size=8)
        y += 14
    pdf.output(str(path))


# Italian month abbreviations for the DATA column (dd mmm yyyy)
_MONTH_IT = {1: "gen", 2: "feb", 3: "mar", 4: "apr", 5: "mag", 6: "giu",
             7: "lug", 8: "ago", 9: "set", 10: "ott", 11: "nov", 12: "dic"}


# --- Trade Republic CSV ---------------------------------------------------------------

_TR_CSV_FIELDS = ["transaction_id", "date", "type", "status", "currency", "amount",
                  "fee", "tax", "name", "description", "counterparty_name",
                  "category", "mcc_code", "asset_class", "isin"]


def write_tr_csv(path: Path, movs: list[TMov], rng: random.Random) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_TR_CSV_FIELDS)
        w.writeheader()
        for t in movs:
            w.writerow({
                "transaction_id": f"{rng.getrandbits(64):016x}",
                "date": t.d.isoformat(),
                "type": t.csv_type,
                "status": "EXECUTED",
                "currency": "EUR",
                "amount": format(t.amount, "f"),
                "fee": format(t.fee, "f") if t.fee else "",
                "tax": format(t.tax, "f") if t.tax else "",
                "name": t.name,
                "description": t.desc_csv,
                "counterparty_name": t.name if t.csv_type == "DEPOSIT" else "",
                "category": t.cat,
                "mcc_code": t.mcc,
                "asset_class": t.asset,
                "isin": t.isin,
            })


# --- other Italian banks (no cashato parser yet) -----------------------------------------
#
# Formats replicated VERBATIM from public, anonymized references (never from
# real data): the BananaAccounting/Italia import-extension testcases (UniCredit,
# BPER, Banca Popolare di Sondrio) and the sources of ofxstatement plugins
# (frankIT/ofxstatement-fineco incl. its test workbook, ecorini/ofxstatement-
# it-banks for Widiba and Webank). They are inputs for FUTURE adapters, so the
# only executable check is that the current detection does NOT misroute them.

_IT_MONTHS_LONG = ["gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                   "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

# (kind, description, amount range)
_OB_KINDS = [
    ("POS", "PAGAMENTO POS ESSELUNGA MILANO", -95, -12),
    ("POS", "PAGAMENTO POS FARMACIA CENTRALE MILANO", -42, -9),
    ("POS", "PAGAMENTO POS Q8 MILANO", -70, -35),
    ("SDD", "ADDEBITO SEPA DD PER FATTURA A VOSTRO CARICO ENEL ENERGIA", -95, -45),
    ("BON_OUT", "DISPOSIZIONE DI BONIFICO ROSSI IMMOBILIARE CANONE", -850, -850),
    ("PRELIEVO", "PRELIEVO ATM MILANO VIA MANZONI", -200, -100),
    ("FEE", "COMMISSIONI E SPESE TENUTA CONTO", -4, -1),
    ("BON_IN", "BONIFICO A VOSTRO FAVORE ACME SRL EMOLUMENTI", 2450, 2490),
]


def _ob_history(rng: random.Random, n: int = 26) -> list[tuple[date, date, str, str, Decimal]]:
    """Six months of generic 2026 movements, NEWEST FIRST like the real exports.

    Descriptions deliberately avoid the word "operazione": several of these
    formats already carry "importo" in their header, and the pair would trip
    Intesa's content detection (["operazione", "importo"]).
    """
    rows = []
    for _ in range(n):
        kind, desc, lo, hi = _OB_KINDS[rng.randrange(len(_OB_KINDS))]
        b = date(2026, rng.randint(1, 6), rng.randint(1, 28))
        v = b + timedelta(days=rng.choice([0, 0, 1]))
        amount = eur(lo) if lo == hi else eur(round(rng.uniform(lo, hi), 2))
        rows.append((b, v, kind, desc, amount))
    rows.sort(key=lambda r: r[0], reverse=True)
    return rows


def _num_it(v: Decimal) -> str:
    return ("-" if v < 0 else "") + fmt_it(v)


def write_unicredit_format1(path: Path, rng: random.Random) -> None:
    """Excel-converted CSV: 2-row preamble, header SPLIT across two rows,
    dd/mm/yyyy + Italian amounts (the UTF-8-BOM testcase variant)."""
    caus = {"POS": "280", "SDD": "198", "BON_OUT": "208", "PRELIEVO": "043",
            "FEE": "016", "BON_IN": "048"}
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Rapporto IT60Y0200801000100000012345 - BIANCHI MARIO", "", "", "", ""])
        w.writerow(["", "", "", "", ""])
        w.writerow(["Data", "", "Descrizione", "EUR", "Caus."])
        w.writerow(["Operaz.", "Valuta", "", "", ""])
        for b, v, kind, desc, amount in _ob_history(rng):
            w.writerow([f"{b:%d/%m/%Y}", f"{v:%d/%m/%Y}", desc, _num_it(amount), caus[kind]])


def write_unicredit_format2(path: Path, rng: random.Random) -> None:
    """Direct CSV export: single header row with a TRAILING semicolon (empty 5th
    column), dd.mm.yyyy, Italian amounts, descriptions padded to 50-char runs."""
    with open(path, "w", newline="", encoding="ascii") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["Data Registrazione", "Data valuta", "Descrizione", "Importo (EUR)", ""])
        for b, v, _kind, desc, amount in _ob_history(rng):
            padded = f"{desc:<50}" + f"{'TRN:11XXXXXXXXX0040873':<50}".rstrip() + " "
            w.writerow([f"{b:%d.%m.%Y}", f"{v:%d.%m.%Y}", padded, _num_it(amount), ""])


def write_bper_xlsx(path: Path, rng: random.Random) -> None:
    """BPER "Smart Web" grid: 11 preamble rows, header at row 12, Italian
    long-form dates, Entrate/Uscite numeric columns, Totale + footer."""
    import openpyxl

    style = {"POS": ("PAGAMENTO POS", "PORTAFOGLIO"), "SDD": ("ADDEBITO", "ADDEBITO SDD/RID"),
             "BON_OUT": ("BONIFICO", "BONIFICO"), "BON_IN": ("BONIFICO", "BONIFICO"),
             "PRELIEVO": ("Prelievo atm", "PRELIEVO"), "FEE": ("CANONE", "CANONE")}
    movs = _ob_history(rng)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Modena, 30 giugno 2026"])
    ws.append([])
    ws.append(["CONTO", "1234567", None, "Saldo disponibile", "4.821,90 EUR"])
    ws.append(["Conto Corrente", "1234567", None, "Saldo contabile", "4.821,90 EUR"])
    ws.append(["IBAN", "IT84M0538701000100000012345", None, "Saldo iniziale", "3.210,74 EUR"])
    ws.append(["Intestatari", "BIANCHI MARIO", None, "Saldo finale", "4.821,90 EUR"])
    ws.append([])
    ws.append([])
    ws.append(["Stai visualizzando i seguenti movimenti:"])
    ws.append([f"Numero Movimenti: {len(movs)}"])
    ws.append(["Intervallo di tempo: dal 1 gennaio 2026 fino al 30 giugno 2026"])
    ws.append(["Data operazione", "Data valuta", "Descrizione", "Entrate", "Uscite",
               "Categoria", "Stato"])

    def d_it(d: date) -> str:
        return f"{d.day} {_IT_MONTHS_LONG[d.month - 1]} {d.year}"

    tot_in = tot_out = D0
    for b, v, kind, _desc, amount in movs:
        descr, cat = style[kind]
        entrate = float(amount) if amount > 0 else None
        uscite = float(amount) if amount < 0 else None
        tot_in, tot_out = tot_in + max(amount, D0), tot_out + min(amount, D0)
        ws.append([d_it(b), d_it(v), descr, entrate, uscite, cat, "Contabilizzato"])
    ws.append([None, None, "Totale", fmt_it(tot_in), _num_it(tot_out)])
    ws.append([])
    ws.append(["Dati Aggiornati al 30 giugno 2026"])
    wb.save(str(path))


def write_bps_csv(path: Path, rng: random.Random) -> None:
    """Popolare di Sondrio "Scrigno" export: every field double-quoted, explicit
    +/- sign, dot decimal, NO thousands separator."""
    causale = {"POS": "PAGAMENTO POS", "SDD": "ADDEBITO DIRETTO", "BON_OUT": "VOSTRA DISPOSIZIONE",
               "BON_IN": "RICEZIONE BONIFICO ISTANTANEO", "PRELIEVO": "PRELIEVO", "FEE": "COMMISSIONI"}
    with open(path, "w", newline="", encoding="ascii") as f:
        w = csv.writer(f, delimiter=";", quoting=csv.QUOTE_ALL)
        w.writerow(["Data", "Valuta", "Causale", "Descrizione", "Importo", "Divisa"])
        for b, v, kind, desc, amount in _ob_history(rng):
            signed = f"{'+' if amount >= 0 else '-'}{abs(amount):.2f}"
            w.writerow([f"{b:%d/%m/%Y}", f"{v:%d/%m/%Y}", causale[kind], desc.title(), signed, "EUR"])


def write_fineco_xlsx(path: Path, rng: random.Random) -> None:
    """Fineco movements workbook: sheet "Movimenti", 6 preamble rows (account id
    in A1 after the exact "Conto Corrente: " prefix), header at row 7,
    Entrate XOR Uscite numeric, a "Totale" footer row."""
    import openpyxl

    descr = {"POS": "VISA DEBIT", "SDD": "SEPA Direct Debit", "BON_OUT": "Bonifico SEPA Italia",
             "BON_IN": "Bonifico SEPA Italia", "PRELIEVO": "Prelievo Bancomat",
             "FEE": "Canone Mensile Conto"}
    movs = _ob_history(rng)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Movimenti"
    ws.append(["Conto Corrente: 1234567"])
    ws.append(["Intestazione Conto Corrente: MARIO BIANCHI"])
    ws.append(["Periodo Dal: 01/01/2026 Al: 30/06/2026"])
    ws.append([])
    ws.append(["Risultati Ricerca"])
    ws.append([])
    ws.append(["Data", "Entrate", "Uscite", "Descrizione", "Descrizione_Completa", "Stato"])
    tot_in = tot_out = D0
    for i, (b, _v, kind, desc, amount) in enumerate(movs):
        stato = "Autorizzato" if i < 2 else "Contabilizzato"
        entrate = float(amount) if amount > 0 else None
        uscite = float(amount) if amount < 0 else None
        tot_in, tot_out = tot_in + max(amount, D0), tot_out + min(amount, D0)
        ws.append([f"{b:%d/%m/%Y}", entrate, uscite, descr[kind], desc.capitalize(), stato])
    ws.append(["Totale", float(tot_in), float(tot_out)])
    wb.save(str(path))


def write_widiba_xlsx(path: Path, rng: random.Random) -> None:
    """Widiba movements list: columns A and F EMPTY (data in B..E and G), account
    number in cell D10, booking date as a raw Excel SERIAL number, a
    "Totale (€)" marker row in the amount column position."""
    import openpyxl

    causale = {"POS": "PAGAMENTO POS", "SDD": "ADDEBITO DIRETTO", "BON_OUT": "BONIFICO SEPA",
               "BON_IN": "BONIFICO SEPA", "PRELIEVO": "PRELIEVO ATM", "FEE": "CANONE"}
    wb = openpyxl.Workbook()
    ws = wb.active
    # "Lista Movimenti" would trip Intesa's ["lista movimenti"] marker: avoid it.
    ws.append([None, "Widiba - Movimenti del conto"])
    for _ in range(7):
        ws.append([])
    ws.append([None, "Numero conto:"])  # row 9: label
    ws.append([None, None, None, "6000123456"])  # row 10: D10 = account number (parser reads [9,3])
    ws.append([])
    ws.append([None, "DATA CONT.", "DATA VAL.", "CAUSALE", "DESCRIZIONE", None, "IMPORTO (€)(€)"])

    def serial(d: date) -> int:
        return (d - date(1899, 12, 30)).days

    for b, v, kind, desc, amount in _ob_history(rng):
        ws.append([None, serial(b), serial(v), causale[kind], desc.capitalize(), None,
                   float(amount)])
    ws.append([None, None, None, None, None, None, "Totale (€)"])
    wb.save(str(path))


def write_webank_xls(path: Path, rng: random.Random) -> None:
    """Webank "movimentiConto.xls": actually an HTML table (the plugin parses it
    with pandas.read_html, decimal=',' thousands='.'). Only the three columns
    the plugin reads by name are confirmed; the real export may carry more."""
    rows = "\n".join(
        f"<tr><td>{b:%d/%m/%Y}</td><td>{desc.title()}</td><td>{_num_it(amount)}</td></tr>"
        for b, _v, _kind, desc, amount in _ob_history(rng)
    )
    path.write_text(
        "<html><body><table>\n"
        "<tr><th>Data Contabile</th><th>Causale / Descrizione</th><th>Importo</th></tr>\n"
        f"{rows}\n</table></body></html>\n",
        encoding="utf-8",
    )


# --- other-bank PDFs (layout contracts from third-party open-source parsers) ---

_DEJAVU = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"


def write_poste_ec_pdf(path: Path, rng: random.Random) -> None:
    """BancoPosta "Estratto Conto" PDF, to the layout contract of
    genbs/poste-italiane-parser (PyMuPDF, coordinate-based).

    Contract highlights: cells are matched by the span's RIGHT edge falling in
    per-column x-windows (data [70,75], valuta [120,125], addebiti [210,230],
    accrediti [300,320], operazione [360,600]); table clip y>=300 on page 1,
    y>=100 after; required metadata areas on page 1 (holder, currency "Euro",
    generated_at, account number digits, IBAN); SALDO INIZIALE/FINALE and
    TOTALE ENTRATE/USCITE meta-rows whose amounts MUST balance; Italian amounts
    WITHOUT a currency symbol (a euro sign would crash its float())."""
    movs = _ob_history(rng)
    opening = eur(1500)
    tot_in = sum((m[4] for m in movs if m[4] > 0), D0)
    tot_out = -sum((m[4] for m in movs if m[4] < 0), D0)
    closing = opening + tot_in - tot_out

    pdf = new_pdf()
    pdf.add_page()
    put(pdf, 40, 60, "Poste Italiane S.p.A. - Patrimonio BancoPosta", size=10, style="B")
    # "ESTRATTO DEL CONTO": the real title phrase "estratto conto" is an Intesa
    # detection marker in cashato; "del" keeps this file out of the wrong parser.
    put(pdf, 40, 80, "ESTRATTO DEL CONTO BANCOPOSTA", size=11, style="B")
    put_right(pdf, 422, 48, "30/06/26", size=7)                  # generated_at [390,40,430,50]
    put(pdf, 392, 78, "Euro", size=8)                            # currency [390,68,420,82]
    put(pdf, 362, 92, HOLDER_GIVEN_FIRST, size=8)                # holder [360,82,500,94]
    put(pdf, 38, 166, "IBAN IT22J0760101600000012345678", size=6)  # iban [36,160,240,168]
    put(pdf, 292, 195, HOLDER_GIVEN_FIRST, size=8)               # customer [290,180,510,220]
    put(pdf, 292, 207, "VIA GARIBALDI 42 20121 MILANO", size=8)
    put(pdf, 152, 214, "1234567890", size=8)                     # account no. [150,200,220,220]
    # column labels ABOVE the page-1 table clip (y>=300): an in-clip header row
    # would be fed to the row parser and crash its date parsing
    put(pdf, 40, 292, "Data / Valuta / Addebiti / Accrediti / Descrizione operazioni", size=7, style="B")

    def row(y: float, d1: str, d2: str, deb: Decimal | None, cred: Decimal | None, op: str) -> None:
        put_right(pdf, 73, y, d1, size=7)                        # data -> right edge in [70,75]
        if d2:
            put_right(pdf, 123, y, d2, size=7)                   # valuta -> [120,125]
        if deb is not None:
            put_right(pdf, 225, y, fmt_it(deb), size=7)          # addebiti -> [210,230]
        if cred is not None:
            put_right(pdf, 315, y, fmt_it(cred), size=7)         # accrediti -> [300,320]
        put(pdf, 365, y, op, size=7, max_x=598)                  # operazione -> [360,600]

    y = 315.0  # page-1 table clip starts at y=300
    row(y, "01/01/26", "", None, opening, "SALDO INIZIALE")
    y += 18  # generous spacing: adjacent rows must land in SEPARATE PyMuPDF blocks
    for b, v, _kind, desc, amount in reversed(movs):  # statement runs oldest-first
        if y > 790:
            pdf.add_page()
            y = 110.0  # later-page clip starts at y=100
        deb = -amount if amount < 0 else None
        cred = amount if amount > 0 else None
        row(y, f"{b:%d/%m/%y}", f"{v:%d/%m/%y}", deb, cred, desc[:52])
        y += 18
    # NO "TOTALE ENTRATE/USCITE" meta-rows: when present the parser compares
    # them to its own float sum WITHOUT rounding, so any accumulation error is
    # fatal; the SALDO FINALE balance check below rounds to 4 decimals instead.
    if y > 790:
        pdf.add_page()
        y = 110.0
    row(y, "30/06/26", "", None, closing, "SALDO FINALE")
    pdf.output(str(path))


def write_hype_pdf(path: Path, rng: random.Random) -> None:
    """Hype (Banca Sella group) movements-list PDF, to the layout contract of
    lorenzogiudici5/ofxstatement-hype: 6 whitespace-separated columns for
    tabula's stream mode (Data Operazione, Data Contabile, Tipologia, Nome,
    Descrizione, Importo), dd/mm/yyyy dates, amount with a LEADING sign and a
    TRAILING euro sign glued on ("-12,34€" — the parser slices both off).
    Tipologia values from the parser's startswith-map."""
    tipologia = {"POS": ("PAGAMENTO", "5411"), "SDD": ("ADDEBITO DIRETTO", ""),
                 "BON_OUT": ("BONIFICO ISTANTANEO", ""), "BON_IN": ("BONIFICO ORDINARIO", ""),
                 "PRELIEVO": ("PAGAMENTO", ""), "FEE": ("CANONE", "")}
    nome = {"POS": "Esselunga", "SDD": "Enel Energia", "BON_OUT": "Rossi Immobiliare",
            "BON_IN": "Acme Srl", "PRELIEVO": "ATM", "FEE": "Hype"}
    pdf = new_pdf()
    pdf.add_font("DejaVu", "", _DEJAVU)
    pdf.add_page()
    put(pdf, 40, 52, "HYPE - Banca Sella S.p.A.", size=10, style="B")
    put(pdf, 40, 70, "Movimenti del conto - MARIO BIANCHI", size=9)
    y = 100.0
    # plausible header labels (= the parser's own column names); tabula reads
    # positions, not names, and the plugin discards any non-date first cell
    for x, label in ((40, "Data Operazione"), (115, "Data Contabile"), (190, "Tipologia"),
                     (290, "Nome"), (360, "Descrizione"), (505, "Importo")):
        pdf.set_font("Helvetica", "B", 7)
        pdf.text(x, y, label)
    y += 16
    for b, v, kind, desc, amount in _ob_history(rng, n=20):
        if y > 790:
            pdf.add_page()
            y = 60.0
        pdf.set_font("Helvetica", "", 7)
        pdf.text(40, y, f"{b:%d/%m/%Y}")
        pdf.text(115, y, f"{v:%d/%m/%Y}")
        pdf.text(190, y, tipologia[kind][0])
        pdf.text(290, y, nome[kind])
        pdf.text(360, y, desc.title()[:28])
        amt = f"{'+' if amount > 0 else '-'}{fmt_it(amount)}€"
        pdf.set_font("DejaVu", "", 7)
        pdf.text(555 - pdf.get_string_width(amt), y, amt)
        y += 15
    pdf.output(str(path))


# The exact operation types g-gg/estratto_ing accepts (closed list), with the
# description shapes its extract_controparte() expects for each.
_ING_ROWS = {
    "BON_IN": ("ACCREDITO BONIFICO",
               "VOSTRO BONIFICO ANAGRAFICA ORDINANTE ACME SRL NOTE: EMOLUMENTI"),
    "BON_OUT": ("VS.DISPOSIZIONE",
                "BONIFICO A FAVORE DI ROSSI IMMOBILIARE BENEF. CANONE LOCAZIONE"),
    "POS": ("PAGAMENTO CARTA", "PAGAMENTO DEL {d} ORE 12:01 PRESSO ESSELUNGA MILANO"),
    "PRELIEVO": ("PRELIEVO CARTA", "PRELIEVO DEL {d} ORE 10:00 PRESSO ATM ING MILANO"),
    "SDD": ("PAGAMENTI DIVERSI",
            "ADDEBITO SDD ENEL ENERGIA CREDITOR ID.IT96ZZZ0000012345678901 ID MANDATO 987654"),
    "FEE": ("BOLLI GOVERNATIVI", "IMPOSTA DI BOLLO SU ESTRATTO"),
}


def write_ing_pdf(path: Path, rng: random.Random) -> None:
    """ING "Conto Arancio" quarterly statement PDF, to the text contract of
    g-gg/estratto_ing (pypdf, line-based state machine): the exact line
    "Estratto conto trimestrale al dd/mm/yyyy", a LISTA MOVIMENTI title, a
    header line containing USCITE repeated on EVERY page, movement lines
    "<data> [<valuta>] <importo> € <TIPO> - <descrizione>" with TIPO from the
    parser's closed list, a RECT_ footer closing every page, and SALDO
    INIZIALE/FINALE that must balance to the cent.

    Keep the text faithful: it is the regression test for the generic
    "estratto conto" / "lista movimenti" detection collision."""
    movs = [m for m in _ob_history(rng, n=45) if date(2026, 4, 1) <= m[0] <= date(2026, 6, 30)]
    opening = eur(1500)
    balance = opening

    pdf = new_pdf()
    pdf.add_font("DejaVu", "", _DEJAVU)

    def line(y: float, text: str, size: float = 8, bold: bool = False) -> None:
        pdf.set_font("Helvetica" if text.isascii() else "DejaVu", "B" if bold and text.isascii() else "", size)
        pdf.text(40, y, text)

    pdf.add_page()
    line(50, "ING BANK N.V. Milan Branch", 10, bold=True)
    line(72, "Estratto conto trimestrale al 30/06/2026", 9)
    line(88, "Conto Corrente Arancio n. 123456 - MARIO BIANCHI", 8)
    line(112, "LISTA MOVIMENTI", 10, bold=True)
    line(128, "DATA OPERAZIONE DATA VALUTA USCITE ENTRATE DESCRIZIONE", 8, bold=True)
    y = 146.0
    rows: list[str] = []
    for b, v, kind, _desc, amount in reversed(movs):  # oldest first
        tipo, desc = _ING_ROWS[kind]
        desc = desc.format(d=f"{b:%d/%m/%y}")
        balance += amount
        rows.append(f"{b:%d/%m/%Y} {v:%d/%m/%Y} {fmt_it(amount)} € {tipo} - {desc}")
    rows.insert(0, f"01/04/2026 {fmt_it(opening)} € SALDO INIZIALE")
    rows.append(f"30/06/2026 {fmt_it(balance)} € SALDO FINALE")
    for i, text in enumerate(rows):
        if y > 770:
            line(786, "Pag. 1", 7)          # lines containing 'Pag.' are skipped
            line(800, "RECT_210320", 6)     # footer marker: closes the page state
            pdf.add_page()
            line(50, "DATA OPERAZIONE DATA VALUTA USCITE ENTRATE DESCRIZIONE", 8, bold=True)
            y = 68.0
        line(y, text, 7)
        y += 14
        if i == len(rows) - 1:
            line(y + 10, "RECT_210320", 6)
    pdf.output(str(path))


def write_other_banks(outdir: Path, rng: random.Random) -> list[tuple[Path, str | None]]:
    """Generate the other-bank files; each entry carries the detection outcome
    cashato's detect_source must produce for it — None for all of them: these
    documents belong to no cashato source and must stay unclaimed."""
    ob = outdir / "other_banks"
    ob.mkdir(parents=True, exist_ok=True)
    writers: list[tuple[Path, object, str | None]] = [
        (ob / "unicredit_movimenti_format1.csv", write_unicredit_format1, None),
        (ob / "unicredit_movimenti_format2.csv", write_unicredit_format2, None),
        (ob / "bper_smart_web_movimenti.xlsx", write_bper_xlsx, None),
        (ob / "popso_scrigno_movimenti.csv", write_bps_csv, None),
        (ob / "fineco_movimenti.xlsx", write_fineco_xlsx, None),
        (ob / "widiba_movimenti.xlsx", write_widiba_xlsx, None),
        (ob / "webank_movimentiConto.xls", write_webank_xls, None),
        (ob / "bancoposta_estratto_conto.pdf", write_poste_ec_pdf, None),
        (ob / "hype_lista_movimenti.pdf", write_hype_pdf, None),
        (ob / "ing_estratto_trimestrale.pdf", write_ing_pdf, None),
    ]
    for path, writer, _expected in writers:
        writer(path, rng)
    return [(p, exp) for p, _w, exp in writers]


# --- verification against the REAL parsers ---------------------------------------------


def verify(outdir: Path, paths: dict[str, list[Path]],
           other: list[tuple[Path, str | None]]) -> bool:
    from cashato.parsers import registry
    from cashato.parsers.detect import detect_source

    ok = True

    def check(cond: bool, msg: str) -> None:
        nonlocal ok
        print(("  PASS  " if cond else "  FAIL  ") + msg)
        ok = ok and cond

    parsed: dict[Path, list] = {}
    for source, files in paths.items():
        for p in files:
            det = detect_source(p)
            check(det == source, f"detect {p.name} -> {det} (expected {source})")
            parsed[p] = registry.ADAPTERS[source](p)
            check(len(parsed[p]) > 0, f"parse  {p.name} -> {len(parsed[p])} transactions")
            keys = [t.natural_key for t in parsed[p]]
            check(len(keys) == len(set(keys)), f"keys   {p.name} -> all natural_keys unique in-file")

    def keyset(p: Path) -> set[str]:
        return {t.natural_key for t in parsed[p]}

    # Revolut: PDF (cash only) ⊂ CSV (cash + interest + crypto)
    rcsv, rpdf = paths["revolut"][0], paths["revolut"][1]
    check(keyset(rpdf) <= keyset(rcsv), "dedup  revolut PDF keys are a subset of CSV keys")
    extra = len(keyset(rcsv) - keyset(rpdf))
    check(extra == len([1 for _ in parsed[rcsv] if _.account in ("revolut_savings_eur", "revolut_crypto")]),
          f"dedup  revolut CSV-only keys = interest+crypto rows ({extra})")

    # Trade Republic: CSV net (amount+fee+tax) == PDF amounts -> identical keys
    tpdf, tcsv = paths["trade_republic"][0], paths["trade_republic"][1]
    check(keyset(tpdf) == keyset(tcsv), "dedup  trade republic PDF and CSV keys identical")

    # Intesa: the 13-month XLSX overlap collapses into the quarterlies
    quarters = [p for p in paths["intesa"] if p.suffix == ".pdf"]
    xlsx = [p for p in paths["intesa"] if p.suffix == ".xlsx"][0]
    qkeys = set().union(*(keyset(p) for p in quarters))
    check(keyset(xlsx) <= qkeys, "dedup  intesa XLSX keys are a subset of the quarterlies'")

    # the deliberate same-day twin survives as TWO distinct keys
    twins = [t for p in quarters for t in parsed[p]
             if t.value_date == date(2025, 3, 14) and t.amount == Decimal("-1.20")]
    check(len(twins) == 2 and twins[0].natural_key != twins[1].natural_key,
          "edge   same-day identical POS pair -> 2 distinct keys (occurrence index)")

    # holder extraction from the PDFs
    for source, p, expected in (("intesa", quarters[0], HOLDER_FAMILY_FIRST),
                                ("revolut", rpdf, HOLDER_GIVEN_FIRST),
                                ("trade_republic", tpdf, HOLDER_GIVEN_FIRST)):
        holder = registry.HOLDER_EXTRACTORS[source](p)
        check(holder == expected, f"holder {p.name} -> {holder!r}")

    # account metadata
    accs = {a.account_id: a for a in registry.ACCOUNT_EXTRACTORS["revolut"](rcsv)}
    check(accs.get("revolut_joint_eur") is not None
          and accs["revolut_joint_eur"].holding_modality == "joint",
          "accts  revolut CSV -> joint account detected as joint")
    intesa_acc = registry.ACCOUNT_EXTRACTORS["intesa"](quarters[0])[0]
    check(intesa_acc.iban == IBAN_INTESA and intesa_acc.product == "XME Conto",
          f"accts  intesa -> product={intesa_acc.product!r} iban ok")

    # internal transfers: opposite legs within 3 days, every month
    all_tx = [t for p in parsed for t in parsed[p]]
    by_key: dict[str, object] = {}
    for t in all_tx:
        by_key.setdefault(t.natural_key, t)
    uniq = list(by_key.values())
    missing = 0
    for m0 in months(START, END):
        legs_out = [t for t in uniq if t.account == "intesa" and t.amount == Decimal("-300.00")
                    and t.value_date.year == m0.year and t.value_date.month == m0.month]
        legs_in = [t for t in uniq if t.account.startswith("revolut_personal")
                   and t.amount == Decimal("300.00")
                   and t.value_date.year == m0.year and t.value_date.month == m0.month]
        if not (legs_out and legs_in):
            missing += 1
    check(missing == 0, "xfer   monthly Intesa->Revolut opposite legs present in every month")

    # other-bank files (future adapters): all must be UNCLAIMED. These two pins
    # are the regression test for the generic Intesa-marker collision, so keep
    # their text faithful.
    for p, expected in other:
        det = detect_source(p)
        check(det == expected, f"detect other_banks/{p.name} -> {det} (expected {expected})")

    # ground-truth export: the expected silver rows after full ingestion
    exp = outdir / "expected_transactions.csv"
    with open(exp, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["natural_key", "source", "account", "value_date", "booking_date",
                    "amount", "currency", "description", "native_category", "mcc"])
        for t in sorted(uniq, key=lambda t: (t.value_date.isoformat(), t.account, str(t.amount))):
            w.writerow([t.natural_key, t.source, t.account, t.value_date, t.booking_date,
                        format(t.amount, "f"), t.currency, t.description,
                        t.native_category or "", t.mcc or ""])
    print(f"\n  {len(all_tx)} parsed rows across {len(parsed)} files "
          f"-> {len(uniq)} unique transactions (dedup removed {len(all_tx) - len(uniq)})")
    print(f"  expected silver rows written to {exp}")
    return ok


# --- main --------------------------------------------------------------------------------


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=Path(__file__).parent, type=Path)
    ap.add_argument("--seed", default=SEED, type=int)
    ap.add_argument("--no-verify", action="store_true")
    args = ap.parse_args()

    outdir: Path = args.out
    outdir.mkdir(parents=True, exist_ok=True)
    rng = random.Random(args.seed)
    gt = build_ground_truth(rng)

    print(f"ground truth: {len(gt.intesa)} intesa, "
          f"{sum(len(v) for v in gt.rev.values())} revolut cash, "
          f"{len(gt.rev_interest)} interest, {len(gt.tr)} trade republic")

    quarters = write_intesa_quarters(outdir, gt.intesa)
    xlsx = outdir / "intesa_lista_operazioni_13m.xlsx"
    write_intesa_xlsx(xlsx, gt.intesa)

    balances = revolut_balances(gt)
    rev_csv = outdir / "revolut_consolidated_statement.csv"
    rev_pdf = outdir / "revolut_consolidated_statement.pdf"
    write_revolut_csv(rev_csv, gt, balances)
    write_revolut_pdf(rev_pdf, gt, balances)

    tr_pdf = outdir / "trade_republic_rendiconto.pdf"
    tr_csv = outdir / "trade_republic_transactions.csv"
    write_tr_pdf(tr_pdf, gt.tr)
    write_tr_csv(tr_csv, gt.tr, rng)

    other = write_other_banks(outdir, rng)

    files = {"intesa": [*quarters, xlsx],
             "revolut": [rev_csv, rev_pdf],
             "trade_republic": [tr_pdf, tr_csv]}
    for p in [p for ps in files.values() for p in ps] + [p for p, _ in other]:
        print(f"  wrote {p} ({p.stat().st_size:,} bytes)")

    if args.no_verify:
        return 0
    print("\nverifying against the real parsers:")
    return 0 if verify(outdir, files, other) else 1


if __name__ == "__main__":
    sys.exit(main())
